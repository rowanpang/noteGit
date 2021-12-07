#!/usr/bin/python

import openpyxl

wb = openpyxl.load_workbook('./lll.xlsx')

print wb.sheetnames
for n in wb.sheetnames:
    ws = wb.get_sheet_by_name(n)
    print ws.max_row
    print ws.max_column
    for rv in ws.values:
        print rv
    print
    for row in ws.rows:
        print row
        for cell in row:
            print cell
            print cell.value
