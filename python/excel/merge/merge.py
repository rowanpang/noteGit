#!/bin/python
# -*- coding: utf-8 -*-

import csv
import openpyxl
import os,sys

srcf = open("/home/pangwz/tmp/src.csv",encoding='gbk')         #csv reader 不会自动转换,默认utf8,需要指明encoding.
srcreader = csv.reader(srcf)

destf = openpyxl.load_workbook('/home/pangwz/tmp/dest.xlsx')
sheet = destf["硬件"]
print(sheet.max_row)
print(sheet.max_column)

i=1
curmax=sheet.max_row            #insert_rows()后再次引用会自动+1;
srcreader.__next__()            #skip header line
notmatch=[]
for row in srcreader:
    vid=row[0]
    adacct=row[1]
    ip=row[2]

    ismatch=0;

    print("try for",vid,ip,adacct)
    for idx in range(3,sheet.max_row+1):
        #print("try match %d" %idx)
        matchcell=sheet.cell(idx,4)                                         #col4: 资产名称
        assetname=matchcell.value.lower()
        if(assetname==adacct):
            ismatch=1
            break

        if (assetname in adacct) or (adacct in assetname):
            print("assetname:",assetname,"adacct",adacct)
            while 1:
                usel=input("Not full matching your select y/n[y]:")
                if len(usel)== 0 or (usel.lower() == "y") or (usel.lower() == "n"):
                    break;
            if (usel.lower() != "n"):
                ismatch=1
                break

    if(ismatch == 0):
        print("no matching")
        notmatch.append(adacct)
    else:
        #print("match %d" %idx)
        newrow=idx+1
        sheet.insert_rows(newrow)
        for srccells in sheet.iter_rows(min_row=idx,max_row=idx):       #获取match 行
            for srccell in srccells:
                sheet.cell(newrow,srccell.column,srccell.value)

        sheet.cell(idx,5,adacct)                                        #old hw terminal
        sheet.cell(newrow,4,vid)                                        #newrow for vdesk 资产名称
        sheet.cell(newrow,5,adacct)                                     #域用户
        sheet.cell(newrow,15,vid)
        sheet.cell(newrow,16,"虚拟桌面,用于查阅IP资料,设计开发")
        sheet.cell(newrow,19,ip)

print("Total no matching %d, %s" %(len(notmatch),notmatch))
destf.save('/home/pangwz/tmp/kkkk.xlsx')
exit()

for row in srcreader:
    vid=row[0]
    user=row[1]
    ip=row[2]

    target=user
    ismatch=0;

    print("try for",vid,ip,user)
    for idx in range(3,curmax+1):
        #print("try match %d" %idx)
        matchcell=sheet.cell(idx,4)
        if(matchcell.value.lower()==target):
            ismatch=1
            #print("match %d" %idx)
            newrow=curmax+i
            i+=1
            sheet.insert_rows(newrow)
            for srccells in sheet.iter_rows(min_row=idx,max_row=idx):       #获取match 行
                for srccell in srccells:
                    sheet.cell(newrow,srccell.column,srccell.value)

            sheet.cell(newrow,4,vid)
            sheet.cell(newrow,14,vid)
            sheet.cell(newrow,18,ip)
            break

    if(ismatch == 0):
        print("not match for user:%s" %user)

destf.save('./kkkk.xlsx')
exit()
rows=sheet.iter_rows(min_row=3,min_col=4,max_col=4)
for row in rows:
    cell=row[0]
    if(cell.value.lower()==target):
        print(cell.row)
        sheet.insert_rows(cell.row+1)
        for srccells in sheet.iter_rows(min_row=cell.row,max_row=cell.row):
            print(srccells)
            for srccell in srccells:
                sheet.cell(cell.row+1,srccell.column,srccell.value)

        sheet.cell(cell.row+1,4,"kkkkk")
        sheet.cell(cell.row+1,18,"192.168.2.111")

destf.save('./kkk.xlsx')
exit()

for i in range(3,5):
    cell=sheet.cell(i,3)
    print(type(cell.value))
    print(cell.encoding)
    print(cell.data_type)
    print(cell.value)
    tmp=cell.value.encode('utf8')
    print(tmp)
    print(type(tmp))
    str=tmp.decode('utf8')
    print(type(str))

exit()
