#!/bin/python
# -*- coding: utf-8 -*-

import csv
import openpyxl
import os

print("庞".encode('utf8'))
srcf = open("./src.csv")
srcreader = csv.reader(srcf)

destf = openpyxl.load_workbook('./dest.xlsx')
destf = openpyxl.load_workbook('./tmp.xlsx')
sheet = destf["Sheet1"]
print(sheet.max_row)
print(sheet.max_column)

rows = sheet.iter_rows(min_row=3,max_col=sheet.max_column)      #same as below
rows = sheet.iter_rows(min_row=3)
rows = sheet.values
print(type(rows))
for row in rows:
    type(row[0])
    print(row[0])
    for cell in row:
        print(cell)

print(sheet['A1'].value)
#for row in sheet.iter_rows('A1:C2'):
#    print(row)

for i in range(1,3):
    cell=sheet.cell(i,2)
    print(cell)
    print(cell.value)

exit()
srcreader.__next__()            #skip header line
for row in srcreader:
   print(row)
   vid=row[0]
   user=row[1]
   ip=row[2]


