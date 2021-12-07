#!/bin/python
# -*- coding: utf-8 -*-

import csv
import openpyxl
import os,sys

#openpyxl 默认已经将xlsx的编码转换为utf8.
#print输出时要结合sys.stdout.encoding 一块看, stdout不一定能encode bytes.
    #UnicodeEncodeError: 'latin-1' codec can't encode characters in position 0-4: ordinal not in range(256)"


print(sys.stdout.encoding)
srcf = open("./src.csv")
srcreader = csv.reader(srcf)

destf = openpyxl.load_workbook('./dest.xlsx')
sheet = destf["硬件"]
print("sheet encoding %s" %sheet.encoding)

print(sys.stdout.encoding)
print(sheet.max_row)
print(sheet.max_column)

for i in range(3,5):
    cell=sheet.cell(i,3)
    print(type(cell.value))
    print(cell.encoding)
    print(cell.data_type)
    print(cell.value)
    tmp=cell.value.encode('utf8')
    print(tmp)
    print(type(tmp))
    str=tmp.decode('utf8')
    print(type(str))

exit()
srcreader.__next__()            #skip header line
for row in srcreader:
   print(row)
   vid=row[0]
   user=row[1]
   ip=row[2]


